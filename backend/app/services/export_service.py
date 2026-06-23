import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from fpdf import FPDF
from datetime import datetime
from typing import List, Dict

class ExportService:
    @staticmethod
    def to_excel(data: List[Dict], report_name: str) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"

        if not data:
            ws.append(["Aviso"])
            ws.append(["No hay datos para este reporte"])
        else:
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([row.get(h, "") for h in headers])

            # Format headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
            thin_side = Side(border_style="thin", color="000000")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def to_pdf(data: List[Dict], report_title: str) -> bytes:
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_font("helvetica", "B", 16)
        
        # Header
        pdf.cell(0, 10, "SERCONIND LTDA.", ln=True, align='L')
        pdf.set_font("helvetica", "I", 8)
        pdf.cell(0, 5, "Ingeniería, Montaje y Obras Civiles", ln=True, align='L')
        pdf.ln(10)
        
        # title
        pdf.set_font("helvetica", "B", 14)
        pdf.cell(0, 10, report_title.upper(), ln=True, align='C')
        pdf.set_font("helvetica", "", 8)
        pdf.cell(0, 5, f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align='R')
        pdf.ln(5)
        
        if not data:
            pdf.cell(0, 10, "No hay datos disponibles para este reporte.", ln=True)
            return pdf.output()

        # Table logic (legacy approach for fpdf without built-in table)
        # However, fpdf2 *has* a table() method which is better.
        # Let's use the table() method if possible.
        
        columns = list(data[0].keys())
        # Convert data to list of lists for fpdf2 table
        table_data = [columns]
        for row in data:
            table_data.append([str(row[col]) for col in columns])

        pdf.set_font("helvetica", size=8)
        with pdf.table(
            borders_layout="HORIZONTAL_LINES",
            cell_fill_color=(240, 240, 240),
            cell_fill_mode="ROWS",
            line_height=6,
            text_align="CENTER"
        ) as table:
            for data_row in table_data:
                row = table.row()
                for datum in data_row:
                    row.cell(datum)

        return bytes(pdf.output())
