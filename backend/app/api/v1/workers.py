from fastapi import APIRouter, Depends, HTTPException, status, Response, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
from app.db.session import get_db
from app.models.core import Employee, UserRole, ProjectAssignment, VacationRequest
from app.schemas.worker import EmployeeCreate, EmployeeOut, EmployeeUpdate, VacationRequestCreate, VacationRequestOut
from app.utils.rut import validate_rut, format_rut
from app.services.pdf_service import ContractService
from app.api.deps import get_current_user, RoleChecker
import os
import io

router = APIRouter()

# Dependencias de rol
# HR_MANAGER, ADMIN y SUPER_ADMIN pueden crear/editar/eliminar
allow_manage_hr = RoleChecker([UserRole.HR_MANAGER, UserRole.ADMIN, UserRole.SUPER_ADMIN])
allow_read_hr = RoleChecker([UserRole.HR_MANAGER, UserRole.PROJECT_MANAGER, UserRole.ADMIN, UserRole.SUPER_ADMIN, UserRole.MANAGEMENT])

# ─── Rutas estáticas primero (deben ir ANTES de /{worker_id}) ────────────────

@router.get("/template-excel")
def download_excel_template():
    import openpyxl
    import io
    from fastapi.responses import Response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Trabajadores"
    
    headers = [
        "Nombres", "Apellidos", "RUT", "Edad", "Cargo", "Sueldo Base", 
        "Fecha Ingreso (AAAA-MM-DD)", "Tipo Contrato (INDEFINIDO/PLAZO_FIJO)", 
        "Fecha Vencimiento Contrato (AAAA-MM-DD)", "Email", "Telefono", "Direccion"
    ]
    ws.append(headers)
    
    ws.append([
        "Juan Carlos", "Perez Gomez", "12.345.678-9", "35", "Maestro Albañil", "850000",
        "2026-01-15", "INDEFINIDO", "", "juan@gmail.com", "+56912345678", "Av San Martin 123"
    ])
    ws.append([
        "Maria Leonor", "Soto Diaz", "15.432.109-8", "28", "Soldador HDPE", "950000",
        "2026-02-01", "PLAZO_FIJO", "2026-08-01", "maria@gmail.com", "+56987654321", "Pasaje El Roble 456"
    ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_trabajadores.xlsx"
        }
    )

@router.post("/import-excel", dependencies=[Depends(allow_manage_hr)])
async def import_workers_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    import openpyxl
    from datetime import datetime
    
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".xlsx", ".xls"]:
        raise HTTPException(status_code=400, detail="Formato de archivo inválido. Debe subir un archivo Excel (.xlsx).")
        
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="El archivo está vacío o no contiene filas de datos.")
            
        data_rows = rows[1:]
        imported_count = 0
        errors = []
        
        for idx, row in enumerate(data_rows, start=2):
            if not row or all(v is None for v in row):
                continue
                
            try:
                first_name = str(row[0]).strip() if row[0] is not None else ""
                last_name = str(row[1]).strip() if row[1] is not None else ""
                rut = str(row[2]).strip() if row[2] is not None else ""
                age_raw = row[3]
                role = str(row[4]).strip() if row[4] is not None else ""
                salary_raw = row[5]
                hire_date_raw = row[6]
                contract_type_raw = str(row[7]).strip().upper() if row[7] is not None else "INDEFINIDO"
                contract_end_raw = row[8]
                email = str(row[9]).strip() if row[9] is not None else ""
                phone = str(row[10]).strip() if row[10] is not None else ""
                address = str(row[11]).strip() if row[11] is not None else ""
                
                if not first_name or not last_name:
                    errors.append(f"Fila {idx}: Nombre y Apellido son obligatorios.")
                    continue
                if not role:
                    errors.append(f"Fila {idx}: El cargo/rol es obligatorio.")
                    continue
                    
                formatted_rut = None
                if rut:
                    if not validate_rut(rut):
                        errors.append(f"Fila {idx}: RUT '{rut}' inválido.")
                        continue
                    formatted_rut = format_rut(rut)
                    existing = db.query(Employee).filter(Employee.rut == formatted_rut).first()
                    if existing:
                        errors.append(f"Fila {idx}: Trabajador con RUT '{rut}' ya existe en el sistema.")
                        continue
                
                age = int(age_raw) if age_raw is not None else None
                salary = int(salary_raw) if salary_raw is not None else 0
                
                hire_date = datetime.utcnow()
                if hire_date_raw:
                    if isinstance(hire_date_raw, datetime):
                        hire_date = hire_date_raw
                    elif isinstance(hire_date_raw, str):
                        try:
                            hire_date = datetime.strptime(hire_date_raw, "%Y-%m-%d")
                        except ValueError:
                            try:
                                hire_date = datetime.strptime(hire_date_raw, "%d/%m/%Y")
                            except ValueError:
                                errors.append(f"Fila {idx}: Formato de Fecha Ingreso inválido.")
                                continue
                
                contract_end_date = None
                if contract_end_raw:
                    if isinstance(contract_end_raw, datetime):
                        contract_end_date = contract_end_raw
                    elif isinstance(contract_end_raw, str):
                        try:
                            contract_end_date = datetime.strptime(contract_end_raw, "%Y-%m-%d")
                        except ValueError:
                            try:
                                contract_end_date = datetime.strptime(contract_end_raw, "%d/%m/%Y")
                            except ValueError:
                                errors.append(f"Fila {idx}: Formato de Fecha Término Contrato inválido.")
                                continue
                
                if contract_type_raw not in ["INDEFINIDO", "PLAZO_FIJO"]:
                    contract_type_raw = "INDEFINIDO"
                
                new_worker = Employee(
                    organization_id=current_user.organization_id,
                    first_name=first_name,
                    last_name=last_name,
                    rut=formatted_rut,
                    age=age,
                    role=role,
                    salary=salary,
                    hire_date=hire_date,
                    contract_type=contract_type_raw,
                    contract_end_date=contract_end_date,
                    email=email,
                    phone=phone,
                    address=address,
                    vacation_balance=15.0
                )
                db.add(new_worker)
                imported_count += 1
            except Exception as row_err:
                errors.append(f"Fila {idx}: Error inesperado - {str(row_err)}")
                
        db.commit()
        return {
            "imported_count": imported_count,
            "errors": errors,
            "success": len(errors) == 0
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error importando Excel: {e}")
        raise HTTPException(status_code=500, detail="Error interno al procesar el archivo Excel.")

# ─── Rutas dinámicas con /{worker_id} ────────────────────────────────────────

@router.get("/{worker_id}/contract", dependencies=[Depends(allow_read_hr)])
def get_worker_contract(
    worker_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    query = db.query(Employee).filter(Employee.id == worker_id)
    if current_user.organization_id:
        query = query.filter(Employee.organization_id == current_user.organization_id)
    worker = query.first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado o no pertenece a su organización")
    
    worker_data = {
        "first_name": worker.first_name,
        "last_name": worker.last_name,
        "rut": worker.rut,
        "role": worker.role,
        "salary": worker.salary,
        "hire_date": worker.hire_date
    }
    
    pdf_content = ContractService.generate_worker_contract(worker_data)
    
    return Response(
        content=pdf_content,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=contrato_{worker.rut}.pdf"
        }
    )

@router.get("/", response_model=List[EmployeeOut], dependencies=[Depends(allow_read_hr)])
def read_workers(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    query = db.query(Employee)
    if current_user.organization_id:
        query = query.filter(Employee.organization_id == current_user.organization_id)
    workers = query.offset(skip).limit(limit).all()
    return workers

@router.post("/", response_model=EmployeeOut, dependencies=[Depends(allow_manage_hr)])
def create_worker(
    worker_in: EmployeeCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    if not validate_rut(worker_in.rut):
        raise HTTPException(status_code=400, detail="RUT inválido")
    
    worker_in.rut = format_rut(worker_in.rut)
    
    # Validar RUT duplicado global o por organización (los RUTs son únicos globales usualmente)
    db_worker = db.query(Employee).filter(Employee.rut == worker_in.rut).first()
    if db_worker:
        raise HTTPException(status_code=400, detail="El trabajador ya está registrado")
    
    # Filtrar solo los campos que existen en la tabla de la BD
    db_keys = {c.key for c in Employee.__table__.columns}
    worker_dict = {k: v for k, v in worker_in.model_dump().items() if k in db_keys}
    new_worker = Employee(**worker_dict)
    new_worker.organization_id = current_user.organization_id
    db.add(new_worker)
    db.commit()
    db.refresh(new_worker)
    
    from app.utils.audit import log_audit
    log_audit(
        db=db,
        user_id=current_user.id,
        action="CREATE",
        table_name="employees",
        record_id=str(new_worker.id),
        new_values=worker_dict
    )
    
    return new_worker

@router.get("/{worker_id}", response_model=EmployeeOut, dependencies=[Depends(allow_read_hr)])
def read_worker(
    worker_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    query = db.query(Employee).filter(Employee.id == worker_id)
    if current_user.organization_id:
        query = query.filter(Employee.organization_id == current_user.organization_id)
    worker = query.first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado o no pertenece a su organización")
    return worker

@router.put("/{worker_id}", response_model=EmployeeOut, dependencies=[Depends(allow_manage_hr)])
def update_worker(
    worker_id: int,
    worker_in: EmployeeUpdate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    query = db.query(Employee).filter(Employee.id == worker_id)
    if current_user.organization_id:
        query = query.filter(Employee.organization_id == current_user.organization_id)
    db_worker = query.first()
    if not db_worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado o no pertenece a su organización")
    
    update_data = worker_in.model_dump(exclude_unset=True)
    
    if "rut" in update_data:
        if not validate_rut(update_data["rut"]):
            raise HTTPException(status_code=400, detail="RUT inválido")
        update_data["rut"] = format_rut(update_data["rut"])
    
    # Capturar valores anteriores antes de modificar
    old_values = {}
    for field in update_data.keys():
        old_values[field] = getattr(db_worker, field)
    
    for field, value in update_data.items():
        setattr(db_worker, field, value)
    
    db.commit()
    db.refresh(db_worker)
    
    from app.utils.audit import log_audit
    log_audit(
        db=db,
        user_id=current_user.id,
        action="UPDATE",
        table_name="employees",
        record_id=str(db_worker.id),
        old_values=old_values,
        new_values=update_data
    )
    
    return db_worker

@router.delete("/{worker_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(allow_manage_hr)])
def delete_worker(
    worker_id: int,
    force: bool = False,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    query = db.query(Employee).filter(Employee.id == worker_id)
    if current_user.organization_id:
        query = query.filter(Employee.organization_id == current_user.organization_id)
    db_worker = query.first()
    if not db_worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado o no pertenece a su organización")
        
    # Verificar si tiene asignaciones activas
    active_assignments = db.query(ProjectAssignment).filter(
        ProjectAssignment.worker_id == worker_id,
        ProjectAssignment.is_active == True
    ).all()
    
    if len(active_assignments) > 0 and not force:
        projects_list = [a.project.name for a in active_assignments]
        raise HTTPException(
            status_code=409,
            detail={
                "code": "ACTIVE_ASSIGNMENTS_WARNING",
                "message": f"El trabajador está asignado a obras activas: {', '.join(projects_list)}.",
                "count": len(active_assignments)
            }
        )
    
    # Liberar asignaciones
    from datetime import datetime
    for assignment in active_assignments:
        assignment.is_active = False
        assignment.unassigned_at = datetime.utcnow()
    
    # Desactivación lógica en lugar de borrado físico
    old_status = getattr(db_worker, "status", "ACTIVE")
    db_worker.status = "INACTIVE"
    db.commit()
    
    from app.utils.audit import log_audit
    log_audit(
        db=db,
        user_id=current_user.id,
        action="DELETE",
        table_name="employees",
        record_id=str(db_worker.id),
        old_values={"status": old_status},
        new_values={"status": "INACTIVE"}
    )
    
    return None



def generate_vacation_document_pdf(employee, request) -> bytes:
    from app.services.pdf_service import BasePDFService
    from datetime import datetime
    pdf = BasePDFService()
    pdf.add_page()
    pdf.set_font("helvetica", size=12)
    
    pdf.set_font("helvetica", 'B', 14)
    pdf.cell(0, 10, "COMPROBANTE DE SOLICITUD DE VACACIONES", ln=True, align='C')
    pdf.ln(10)
    
    pdf.set_font("helvetica", size=11)
    pdf.cell(0, 8, f"Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y')}", ln=True)
    pdf.ln(5)
    
    pdf.set_font("helvetica", 'B', 11)
    pdf.cell(0, 8, "DATOS DEL TRABAJADOR", ln=True)
    pdf.set_font("helvetica", size=11)
    pdf.cell(0, 6, f"Nombre Completo: {employee.first_name} {employee.last_name}", ln=True)
    pdf.cell(0, 6, f"RUT: {employee.rut or 'Sin especificar'}", ln=True)
    pdf.cell(0, 6, f"Cargo: {employee.role}", ln=True)
    pdf.ln(5)
    
    pdf.set_font("helvetica", 'B', 11)
    pdf.cell(0, 8, "DETALLE DEL PERIODO DE VACACIONES", ln=True)
    pdf.set_font("helvetica", size=11)
    pdf.cell(0, 6, f"Desde: {request.start_date.strftime('%d/%m/%Y')}", ln=True)
    pdf.cell(0, 6, f"Hasta: {request.end_date.strftime('%d/%m/%Y')}", ln=True)
    pdf.cell(0, 6, f"Días Solicitados: {request.days_requested} días hábiles", ln=True)
    pdf.cell(0, 6, f"Saldo Disponible previo a Solicitud: {employee.vacation_balance} días", ln=True)
    pdf.ln(10)
    
    pdf.multi_cell(0, 6, 
        "Por medio del presente comprobante, se autoriza y deja constancia del periodo de descanso legal del trabajador. "
        "El trabajador se compromete a reincorporarse a sus labores operacionales el día hábil inmediato siguiente a la fecha de término."
    )
    pdf.ln(25)
    
    y = pdf.get_y()
    pdf.line(20, y, 90, y)
    pdf.line(120, y, 190, y)
    pdf.set_y(y + 2)
    pdf.set_x(20)
    pdf.cell(70, 5, "Firma Trabajador", align='C')
    pdf.set_x(120)
    pdf.cell(70, 5, "Firma Representante / Autorizador", align='C')
    
    return pdf.output()

@router.post("/vacations/request", response_model=VacationRequestOut, dependencies=[Depends(allow_read_hr)])
def create_vacation_request(
    req_in: VacationRequestCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    query = db.query(Employee).filter(Employee.id == req_in.employee_id)
    if current_user.organization_id:
        query = query.filter(Employee.organization_id == current_user.organization_id)
    worker = query.first()
    if not worker:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado en su organización")
        
    if worker.vacation_balance < req_in.days_requested:
        raise HTTPException(
            status_code=400, 
            detail=f"Días insuficientes. El trabajador tiene {worker.vacation_balance} días disponibles y solicita {req_in.days_requested}."
        )
        
    new_req = VacationRequest(
        employee_id=req_in.employee_id,
        start_date=req_in.start_date,
        end_date=req_in.end_date,
        days_requested=req_in.days_requested,
        status="PENDING_APPROVAL"
    )
    db.add(new_req)
    
    from app.services.notification_service import NotificationService
    from app.models.core import NotificationType, NotificationPriority
    NotificationService._create_notification_if_not_exists(
        db,
        NotificationType.VACATION_REQUEST,
        worker.id,
        title="Nueva Solicitud de Vacaciones",
        message=f"El trabajador {worker.first_name} {worker.last_name} solicita {req_in.days_requested} días de vacaciones desde el {req_in.start_date.strftime('%Y-%m-%d')}.",
        priority=NotificationPriority.INFO,
        link="/workers"
    )
    
    db.commit()
    db.refresh(new_req)
    return new_req

@router.patch("/vacations/request/{request_id}/approve", response_model=VacationRequestOut)
def approve_vacation_request(
    request_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    from app.models.core import UserRole
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.HR_MANAGER]:
        raise HTTPException(status_code=403, detail="No tiene permisos para autorizar vacaciones.")
        
    query = db.query(VacationRequest).join(Employee).filter(VacationRequest.id == request_id)
    if current_user.organization_id:
        query = query.filter(Employee.organization_id == current_user.organization_id)
    req = query.first()
    if not req:
        raise HTTPException(status_code=404, detail="Solicitud de vacaciones no encontrada o no pertenece a su organización")
        
    if req.status != "PENDING_APPROVAL":
        raise HTTPException(status_code=400, detail=f"La solicitud no se encuentra pendiente. Estado actual: {req.status}")
        
    req.status = "APPROVED"
    req.approved_by = current_user.id
    
    from app.services.notification_service import NotificationService
    from app.models.core import NotificationType, NotificationPriority
    NotificationService._create_notification_if_not_exists(
        db,
        NotificationType.VACATION_APPROVED,
        req.employee_id,
        title="Deducción de Vacaciones Pendiente",
        message=f"Las vacaciones de {req.employee.first_name} {req.employee.last_name} han sido AUTORIZADAS por {current_user.full_name}. Proceda con la firma del documento y la rebaja de {req.days_requested} días.",
        priority=NotificationPriority.WARNING,
        link="/workers"
    )
    
    try:
        pdf_bytes = generate_vacation_document_pdf(req.employee, req)
        os.makedirs("uploads/documents", exist_ok=True)
        unique_filename = f"solicitud_vacaciones_{req.id[:8]}.pdf"
        file_path = f"uploads/documents/{unique_filename}"
        with open(file_path, "wb") as f:
            f.write(pdf_bytes)
        req.document_path = f"/uploads/documents/{unique_filename}"
        
        from app.models.core import Document
        db_doc = Document(
            organization_id=req.employee.organization_id,
            title=f"Solicitud Vacaciones Autorizada - {req.employee.first_name}",
            file_path=req.document_path,
            file_type="application/pdf",
            file_size=len(pdf_bytes),
            category="vacaciones",
            employee_id=req.employee_id,
            created_by=current_user.id
        )
        db.add(db_doc)
    except Exception as pdf_err:
        print(f"Error generating vacation PDF: {pdf_err}")
        
    db.commit()
    db.refresh(req)
    return req

@router.patch("/vacations/request/{request_id}/rebate", response_model=VacationRequestOut)
def rebate_vacation_request(
    request_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    from app.models.core import UserRole
    if current_user.role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.HR_MANAGER]:
        raise HTTPException(status_code=403, detail="No tiene permisos para efectuar la rebaja de vacaciones.")
        
    query = db.query(VacationRequest).join(Employee).filter(VacationRequest.id == request_id)
    if current_user.organization_id:
        query = query.filter(Employee.organization_id == current_user.organization_id)
    req = query.first()
    if not req:
        raise HTTPException(status_code=404, detail="Solicitud de vacaciones no encontrada o no pertenece a su organización")
        
    if req.status != "APPROVED":
        raise HTTPException(status_code=400, detail="La solicitud debe estar AUTORIZADA previamente para efectuar la rebaja.")
        
    if not req.is_signed:
        raise HTTPException(status_code=400, detail="El documento debe estar firmado por el trabajador antes de realizar la rebaja.")
        
    employee = req.employee
    if employee.vacation_balance < req.days_requested:
        raise HTTPException(status_code=400, detail=f"El saldo actual del trabajador ({employee.vacation_balance} días) es menor al solicitado.")
        
    employee.vacation_balance = employee.vacation_balance - req.days_requested
    req.status = "REBATED"
    req.rebated_by = current_user.id
    
    db.commit()
    db.refresh(req)
    return req

@router.post("/vacations/request/{request_id}/upload-document")
async def upload_signed_vacation_doc(
    request_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    from app.models.core import Document
    query = db.query(VacationRequest).join(Employee).filter(VacationRequest.id == request_id)
    if current_user.organization_id:
        query = query.filter(Employee.organization_id == current_user.organization_id)
    req = query.first()
    if not req:
        raise HTTPException(status_code=404, detail="Solicitud de vacaciones no encontrada o no pertenece a su organización")
        
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".pdf", ".jpg", ".jpeg", ".png"]:
        raise HTTPException(status_code=400, detail="Formato de archivo no soportado. Suba un archivo PDF o Imagen.")
        
    contents = await file.read()
    
    unique_filename = f"signed_vacation_{req.id[:8]}{ext}"
    file_path = f"uploads/documents/{unique_filename}"
    
    with open(file_path, "wb") as f:
        f.write(contents)
        
    req.document_path = f"/uploads/documents/{unique_filename}"
    req.is_signed = True
    
    db_doc = Document(
        organization_id=req.employee.organization_id,
        title=f"Documento de Vacaciones Firmado - {req.employee.first_name}",
        file_path=req.document_path,
        file_type=file.content_type,
        file_size=len(contents),
        category="vacaciones",
        employee_id=req.employee_id,
        created_by=current_user.id
    )
    db.add(db_doc)
    db.commit()
    
    return {"message": "Documento firmado subido exitosamente", "document_path": req.document_path}

@router.get("/vacations/requests", response_model=List[VacationRequestOut], dependencies=[Depends(allow_read_hr)])
def list_vacation_requests(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    query = db.query(VacationRequest).join(Employee)
    if current_user.organization_id:
        query = query.filter(Employee.organization_id == current_user.organization_id)
    return query.order_by(VacationRequest.created_at.desc()).all()
